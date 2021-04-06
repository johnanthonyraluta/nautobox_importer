import sys
import re
import os
import xlrd
import openpyxl
from openpyxl import load_workbook

class xlsx_handler:

    def __init__(self, filename):
        self.filename=filename
    
    def read(self,worksheet_name):
        workbook=load_workbook(self.filename)
        sheet = workbook.get_sheet_by_name(worksheet_name)
        dict_list = []
        keys = [cell.value for cell in sheet[2]]
        for row in sheet.iter_rows(min_row=3):
            d = {}
            skip=False
            c=0
            for cell in row:
                if not cell.font.strike:
                    if keys[c]!=None:
                        d[keys[c].lower()]=cell.value
                else:
                    skip=True
                c+=1
            if not skip:
                dict_list.append(d)
        return dict_list

    def read_all(self,worksheet_name):
        workbook=load_workbook(self.filename)
        sheet = workbook.get_sheet_by_name(worksheet_name)
        dict_list = []
        keys = [cell.value for cell in sheet[1]]
        for row in sheet.iter_rows(min_row=2):
            d = {}
            skip=False
            c=0
            for cell in row:
                d['row']=cell.row
                if not cell.font.strike:
                    if keys[c]!=None:
                        d[keys[c].lower()]=cell.value
                else:
                    skip=True
                c+=1
            if not skip:
                dict_list.append(d)
        return dict_list

    #def create_workbook(self,sheetname,output_folder,data):
    #    ws = wb.create_sheet(sheetname)


if __name__ == "__main__":
    import argparse
    from pprint import pprint
    parser = argparse.ArgumentParser()
    parser.add_argument('-f','--file', required=True, help='File')
    parser.add_argument('-w','--worksheet', required=True, help='WorkSheet name')
    args = parser.parse_args()
    xl=xlsx_handler(args.file)
    d=xl.read( args.worksheet )
    pprint(d)
    #print(len(d))
